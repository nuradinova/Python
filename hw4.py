import openpyxl


# workbook = openpyxl.load_workbook("data/data1.xlsx")
# worksheet = workbook.active
# ext_matrix = []
#
# for row in range(1, worksheet.max_row+1):
#     int_matrix = []
#     for column in range(1, worksheet.max_column+1):
#         int_matrix.append(worksheet.cell(row, column).value)
#     ext_matrix.append(int_matrix)
#print(ext_matrix)



# ВТОРОЕ ЗАДАНИЕ
workbook = openpyxl.load_workbook("data/data2.xlsx")
worksheet = workbook.active
ext_matrix = []

for column in range(1, worksheet.max_column+1):
    int_matrix = []
    for row in range(1, worksheet.max_row+1):
        int_matrix.append(worksheet.cell(row, column).value)
    ext_matrix.append(int_matrix)



